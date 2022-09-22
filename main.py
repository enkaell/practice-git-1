# Import load_workbook module from openpyxl
from openpyxl import load_workbook
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import time

SIMA_LAND_TOKEN = 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2NjExODQyNTAsImlhdCI6MTY2MDU3OTQ1MCwianRpIjoyNzA3MjQxLCJuYmYiOjE2NjA1Nzk0NTB9.Ho432tb4FaE-e6aM5izK2mPwm7LrkoPnzcqmGuGIQck'
wb = load_workbook('./assortment_business_8627427_31-07-2022.xlsx')
sheet = wb['Ассортимент']
session = requests.Session()
retry = Retry(connect=3, backoff_factor=0.5)
adapter = HTTPAdapter(max_retries=retry)
j = 0

start_time = int(time.time())
for i in sheet['C'][4:]:
    response = requests.get(
        f'https://www.sima-land.ru/api/v5/item/{i.value}',
        headers={
            'accept': 'application/json',
            'X-Api-Key': SIMA_LAND_TOKEN,
            'Authorization': SIMA_LAND_TOKEN,
        },

        params={
            'view': 'brief',
            'by_sid': 'false',
        },
        timeout=3
    )
    sheet[i.coordinate].value = response.json()['sid']

wb.save('yandex.xlsx')